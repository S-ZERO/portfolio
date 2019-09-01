from selenium import webdriver

import openpyxl
from openpyxl import load_workbook

from common import aguse_reserach_module as aguse
from forIP import mxtoolbox_blacklist_module as mxtoolbox
from forIP import owner_search_IP_module as ownerIP

search = []

aguse_results = []
mxtoolbox_results = []
owner_IP_results = []
xforce_category = []
xforce_time = []
xforce_risk = []
xforce_IP = []

#結果エクセルの見栄え設定(ヘッダーの色、枠線)
def all_result_writer(write_result_list, col_num, target_wb):
    headers = ["IP", "aguse", "mxtoolbox", "owner"]
    fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor='228B22', bgColor='228B22')
    font = openpyxl.styles.fonts.Font(color = 'FFFFFF', bold = True )
    border_line = openpyxl.styles.borders.Side(style='thin', color = '000000')
    border = openpyxl.styles.borders.Border(top=border_line, bottom=border_line, left=border_line, right=border_line)
    print(fill)
    if not "all_result" in target_wb.get_sheet_names():
        all_result_sheet = target_wb.create_sheet(title = "all_result")
        for col, header in enumerate(headers):
            all_result_sheet.cell(row = 1, column = col+1, value = header).fill = fill
            all_result_sheet.cell(row = 1, column = col+1).font = font
            all_result_sheet.cell(row = 1, column = col+1).border = border
    else:
        all_result_sheet = target_wb.get_sheet_by_name("all_result")

    for num, write_value in enumerate(write_result_list):
        all_result_sheet.cell(row = num+2, column = col_num+1, value = write_value)
        all_result_sheet.cell(row = num+2, column = col_num+1).border = border
    
    wb.save(r"C:\addr_research_portfolio\result\result_IP.xlsx")

driver = webdriver.Chrome(r'C:\addr_research_portfolio\common\chromedriver.exe')  

print("読み込ませるファイルを選択")
target_file = input(">> ")
with open(target_file, encoding='utf-8') as f:

    for rows in f:
        row = rows.rstrip('\n\n')
        search.append(row)

for ip in search:
    #------aguseの処理--------------------------------
    for roop in range(0,5):
        aguse_result = aguse.main_move(ip,driver)
        if aguse_result == "safe" or aguse_result == "caution":
            break
    aguse_results.append(aguse_result)
    print(aguse_results)
    
    #------mxtoolboxの処理--------------------------------
    for roop in range(0,5):
       mxtoolbox_result = mxtoolbox.main_move(ip,driver)
       if not mxtoolbox_result == None and int(mxtoolbox_result[1]) < 5:
           break
    mxtoolbox_results.append(mxtoolbox_result[0])
    print(mxtoolbox_results)
    
    #------ownerの処理--------------------------------
    for roop in range(0,5):
       owner_IP_result = ownerIP.main_move(ip,driver)
       if not owner_IP_result == None and owner_IP_result != "Web接続制限":
           break
    owner_IP_results.append(owner_IP_result)
    print(owner_IP_results)

wb = openpyxl.Workbook()
all_results_list = [search, aguse_results, mxtoolbox_results, owner_IP_results]
for num, results_list in enumerate(all_results_list):
    all_result_writer(results_list, num, wb)

print("完了")
driver.close()